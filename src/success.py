#!/usr/bin/python
import os
import sys
import getopt
import matplotlib
import matplotlib.pyplot as plt
import matplotlib.patches as pat
import pytablewriter as ptw
from matplotlib import colors
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, Alignment
from openpyxl.formatting.rule import CellIsRule
import numpy as np
from tqdm import tqdm
from tabulate import tabulate
from collections import OrderedDict
import matplotlib.gridspec as gridspec
from scipy import stats
from CsvDatabase import *
from StatsFilter import *

HELP_MSG = 'main.py -d <domain> -p <planner> -f <fusion-ratio> -g <gridcolor>\nplanner choices:\n\ttfddownward\n\tcolin2\ndomain choices:\n\tblocks_world\n\tfirst_response'
EXCEL_TABLE	= 'xls/success'

domain = 'blocks_world'
planner = 'tfddownward'
fusion_ratio = '0.50'
grid = 'w'
try:
	opts, args = getopt.getopt(sys.argv[1:],'hd:p:f:g:',['domain=','planner=','fusion-ratio=','gridcolor='])
except getopt.GetoptError:
	print HELP_MSG
	sys.exit(2)
for opt, arg in opts:
	if opt in ('-h', '--help'):
		print HELP_MSG
		sys.exit()
	elif opt in ('-p', '--planner'):
		planner = arg
	elif opt in ('-d', '--domain'):
		domain = arg
	elif opt in ('-f', '--fusion-ratio'):
		fusion_ratio = arg
	elif opt in ('-g', '--gridcolor'):
		grid = arg
print 'Domain: %s' % domain
print 'Planner: %s' % planner
print 'Fusion Ratio: %s' % fusion_ratio
print 'Grid: %s' % grid

COLORCODE			= OrderedDict([
							(0,		OrderedDict([	('short', 'Success'),	('code', '0'),		('color', (0.000, 0.447, 0.741))	])), # Blue
							(1,		OrderedDict([	('short', 'Nonexec'),	('code', '1'),		('color', (0.850, 0.325, 0.098))	])), # Tomato
							(124,	OrderedDict([	('short', 'Time Fail'),	('code', '124'),	('color', (0.929, 0.694, 0.125))	])), # Orange
							(134,	OrderedDict([	('short', 'Mem Fail'),	('code', '134'),	('color', (0.929, 0.894, 0.325))	])), # Yellow
					])


TOOLS				= OrderedDict([
							('Object',				OrderedDict([	('reg', 'O'),		('tex',r'\textbf{O}'),		('color', (1.000, 0.000, 0.000)), ('marker', (3,0,0)),		('plt-loc', (0,0))	])), # Red
							('ObjectTime',			OrderedDict([	('reg', 'OT'),		('tex',r'\textbf{OT}'),		('color', (1.000, 0.500, 0.500)), ('marker', (3,0,180)),	('plt-loc', (1,0))	])), # Pink
							('Action',				OrderedDict([	('reg', 'A'),		('tex',r'\textbf{A}'),		('color', (0.000, 0.500, 0.000)), ('marker', (4,0,0)),		('plt-loc', (0,1))	])), # Green
							('ActionTime',			OrderedDict([	('reg', 'AT'),		('tex',r'\textbf{AT}'),		('color', (0.500, 1.000, 0.500)), ('marker', (4,0,45)),		('plt-loc', (1,1))	])), # Light Green
							('ActionObject',		OrderedDict([	('reg', 'AO'),		('tex',r'\textbf{AO}'),		('color', (0.000, 0.000, 1.000)), ('marker', (5,0,0)),		('plt-loc', (0,2))	])), # Blue
							('ActionObjectTime',	OrderedDict([	('reg', 'AOT'),		('tex',r'\textbf{AOT}'),	('color', (0.000, 0.750, 1.000)), ('marker', (5,0,180)),	('plt-loc', (1,2))	])), # Light Blue
							('CoalitionSimilarity',	OrderedDict([	('reg', 'CS'),		('tex','CS'),				('color', (0.500, 0.000, 0.500)), ('marker', (6,0,0)),		('plt-loc', (0,3))	])), # Purple
							('CoalitionAssistance',	OrderedDict([	('reg', 'CA'),		('tex','CA'),				('color', (1.000, 0.500, 1.000)), ('marker', (6,0,30)),		('plt-loc', (1,3))	])), # Light Purple
							('CFP',					OrderedDict([	('reg', 'CFP'),		('tex','CFP'),				('color', (0.000, 0.000, 0.000)), ('marker', (7,0,0)),		('plt-loc', (0,4))	])), # Black
							('PA',					OrderedDict([	('reg', 'PA'),		('tex','PA'),				('color', (0.500, 0.500, 0.500)), ('marker', (7,0,180)),	('plt-loc', (1,4))	])), # Grey
					])


if domain == 'first_response':
	del TOOLS['PA']


PLANNER_DOM			= {'first_response': 'First Response', 'blocks_world': 'Blocks World', 'colin2': 'COLIN', 'tfddownward': 'TFD'}
GROUP_TITLES		= {'0.25':'a', '0.50':'b', '0.75':'c'}

# Labels
file_header			= ['Domain','Problem','CFA','Planner','Tool','Fusion Ratio', 'Planning Results (%)', 'Makespan (s)', 'Number of Actions', 'Processing Time (s)', 'Memory Usage (GB)']

LINEWIDTH			= 0.1
BAR_FILL			= 0.60
FONT_SIZE			= 6
FONT_FAMILY			= 'serif'
COL_PAD				= 5
CSPACING			= 1

# File names
GATHER_DATA_SCRIPT	= "scripts/gather_data.sh"
RAW_STATS			= "csv/stats.csv"
FILTERED_STATS		= "csv/stats_filtered.csv"
PROB_PLOT_NAME		= "plots/prob"
# PLOT_FORMATS		= ['pdf', 'eps', 'svg']
PLOT_FORMATS		= ['svg']
PROBL_FORMAT		= 'P%02dC%02d'

def create_sheet(wb, title, width):

	ws = wb.create_sheet(title=title)

	# Column Number (Coalition)
	for r, row in enumerate(reversed(ws['C2:L2'])):
		for c,cell in enumerate(row):
			cell.value = c+1
			cell.font = Font(bold=True)

	# Row Number (Problem)
	for r, row in enumerate(reversed(ws['B3:B12'])):
		for c,cell in enumerate(row):
			cell.value = r+1
			cell.font = Font(bold=True)

	# Coalition and Problem Labels
	ws['C1'] = 'Coalition'
	ws['C1'].alignment=Alignment(horizontal='center',vertical='center')
	ws['C1'].font=Font(bold=True)
	ws['A3'] = 'Problem'
	ws['A3'].alignment=Alignment(horizontal='center',vertical='center',text_rotation=90)
	ws['A3'].font=Font(bold=True)
	ws.merge_cells('C1:L1')
	ws.merge_cells('A3:A12')

	# Sum over each row (Problem)
	ws.conditional_formatting.add('M3:M12', CellIsRule(operator='greaterThan', stopIfTrue=True, formula=['9'], font=Font(bold=True)))
	for row in ws['M3:M12']:
		for c in row:
			c.value = '=SUM(C%d:L%d)' % (c.row,c.row)

	# Sum over each column (Coalition)
	ws.conditional_formatting.add('C13:L13', CellIsRule(operator='greaterThan', stopIfTrue=True, formula=['9'], font=Font(bold=True)))
	for row in ws['C13:M13']:
		for c in row:
			c.value = '=SUM(%s3:%s12)' % (c.column,c.column)

	# Column Widths
	for col in ws.columns:
		ws.column_dimensions[col[0].column].width = width

	return ws

def generate_excel(data,filename):
	print 'Storing data to %s ...' % filename
	wb = Workbook()
	del wb['Sheet']


	# Main data
	for t in data:
		ws = create_sheet(wb=wb, title=TOOLS[t]['reg'], width=3.0)
		for r, row in enumerate(reversed(ws['C3:L12'])):
			for c,cell in enumerate(row):
				problem = PROBL_FORMAT%(r+1,c+1)
				cell.value = int(data[t][problem])!=0

	# All Tools Failed
	ws = create_sheet(wb=wb, title='All Failed', width=3.0)
	for r, row in enumerate(reversed(ws['C3:L12'])):
		for c,cell in enumerate(row):
			problem = PROBL_FORMAT%(r+1,c+1)
			cell.value = sum([int(data[t][problem])==0 for t in data])==0

	# All Tools Succeeded
	ws = create_sheet(wb=wb, title='All Succeeded', width=3.0)
	for r, row in enumerate(reversed(ws['C3:L12'])):
		for c,cell in enumerate(row):
			problem = PROBL_FORMAT%(r+1,c+1)
			cell.value = sum([int(data[t][problem])>0 for t in data])==0

	# At least some tools failed
	ws = create_sheet(wb=wb, title='Some Failed', width=3.0)
	for r, row in enumerate(reversed(ws['C3:L12'])):
		for c,cell in enumerate(row):
			problem = PROBL_FORMAT%(r+1,c+1)
			cell.value = sum([int(data[t][problem])>0 for t in data])>0

	# At least some tools succeeded
	ws = create_sheet(wb=wb, title='Some Succeeded', width=3.0)
	for r, row in enumerate(reversed(ws['C3:L12'])):
		for c,cell in enumerate(row):
			problem = PROBL_FORMAT%(r+1,c+1)
			cell.value = sum([int(data[t][problem])==0 for t in data])>0

	wb.save(filename)

def generate_plot(data):

	plt.figure(frameon=True)
	f, axes = plt.subplots(2,5, sharex='all', sharey='all', squeeze=True)
	plt.subplots_adjust(hspace=0.15, wspace=-0.20, bottom=0.15, top=0.88, left=0.00, right=1.00)
	plt.gcf().set_size_inches(4.5,1.8)
	matplotlib.rcParams.update({'font.size':			FONT_SIZE})
	matplotlib.rcParams.update({'font.family':			FONT_FAMILY})
	matplotlib.rcParams.update({'axes.linewidth':		0.2})
	matplotlib.rcParams.update({'xtick.major.width':	0.2})
	matplotlib.rcParams.update({'ytick.major.width':	0.2})
	matplotlib.rc('text', usetex=True)


	if domain == 'first_response':
		axes[1,4].axis('off')

	if fusion_ratio == '0.25':
		labels = [COLORCODE[k]['short'] for k in COLORCODE]
		patches = [pat.Patch(facecolor=v['color'], edgecolor=grid, label=k, linewidth=LINEWIDTH) for k,v in COLORCODE.items()]
		axes[0,2].legend(patches, labels, loc='center', frameon=False, ncol=4, bbox_to_anchor=(0.5, 1.25), fontsize=FONT_SIZE)

	for t in data:

		# Plot location indexes
		px = TOOLS[t]['plt-loc'][0]
		py = TOOLS[t]['plt-loc'][1]

		# Plot handle
		ax = axes[px,py]

		# Plot
		for pidx,p in enumerate(range(10)):
			for cidx,c in enumerate(range(10)):
				code = PROBL_FORMAT%(p+1,c+1)
				error_code = int(data[t][code])
				ax.add_patch(pat.Rectangle((cidx-0.5,pidx-0.5), 1.0, 1.0, edgecolor=grid, facecolor=COLORCODE[error_code]['color'], linewidth=LINEWIDTH))

		# Plot Title
		tool_title = TOOLS[t]['tex']
		# if t not in ['PA','CFP']:
		# 	tool_title += ' ($f_{max} = %s$)' % fusion_ratio
		ax.set_title(tool_title, fontsize=FONT_SIZE, position=(0.5,0.9))

		# Square Aspect Ratio
		ax.set_adjustable('box-forced')
		ax.set_aspect(1)
		
		# Major tick indexes
		ax.set_xticks(np.arange(0,10,9.0));
		ax.set_yticks(np.arange(0,10,9.0));

		# Labels for major ticks
		ax.set_xticklabels(np.arange(1, 11, 9));
		ax.set_yticklabels(np.arange(1, 11, 9));

		# Minor ticks
		ax.set_xticks(np.arange(-.45, 10, 1), minor=True);
		ax.set_yticks(np.arange(-.43, 10, 1), minor=True);

		# Hidding tick marks
		ax.xaxis.set_ticks_position('none') 
		ax.yaxis.set_ticks_position('none')

		# Axis labels
		if px == 1:
			ax.set(xlabel='Coalition')
		if py == 0:
			ax.set(ylabel='Problem')
		ax.xaxis.labelpad = -5
		ax.yaxis.labelpad = -7

		for s in ax.spines.values():
			s.update({
					'edgecolor': grid,
					'linewidth': LINEWIDTH,
				})

	plt.suptitle('%s) %s with %s problem-wise results for $f_{max} = %s$.' % (GROUP_TITLES[fusion_ratio], PLANNER_DOM[domain], PLANNER_DOM[planner], fusion_ratio), x=0.30, y=0.05, fontsize=FONT_SIZE)

	for f in PLOT_FORMATS:
		plt.savefig('-'.join([PROB_PLOT_NAME,domain,planner,fusion_ratio])+'.'+f)

# Gathering data
print 'Gathering data ...'
os.system(GATHER_DATA_SCRIPT)

# Filtering CSV file
print 'Filtering CSV file ...'
StatsFilter.filter(RAW_STATS,FILTERED_STATS,file_header)

# Creating database
print 'Creating database ...'
db = CsvDatabase(FILTERED_STATS)

# For each tool
print 'Processing ...'
data = OrderedDict()
for t in TOOLS.keys():
	query_all	= db.query([('Domain',domain),('Planner',planner),('Tool',t),('Fusion Ratio',['0.00',fusion_ratio])])
	problems	= db.select(['Problem','Planning Results (%)'], query_all)
	data[t]		= OrderedDict(problems)


# generate_plot(data)


# Excel Spreadsheet
print 'Excel Spreadsheet ...'
generate_excel(data,'-'.join([EXCEL_TABLE,domain,planner,fusion_ratio])+'.xlsx')
